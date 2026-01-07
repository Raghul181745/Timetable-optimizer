from flask import Flask, render_template, request, redirect, url_for, g, flash, session, make_response
import os
import sqlite3
import csv
import io

app = Flask(__name__)
app.secret_key = "secret123"

# Database setup
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_DIR = os.path.join(BASE_DIR, "Timetable.db")
os.makedirs(DB_DIR, exist_ok=True)
DB_PATH = os.path.join(DB_DIR, "timetable.db")

def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timetable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            staff_name TEXT NOT NULL,
            department TEXT NOT NULL,
            year TEXT NOT NULL,
            semester TEXT NOT NULL,
            subject TEXT NOT NULL,
            day TEXT NOT NULL,
            time TEXT NOT NULL
        )
    ''')
    # Attempt to add the 'year' column if it doesn't exist (for existing databases)
    try:
        cursor.execute("ALTER TABLE timetable ADD COLUMN year TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    # Attempt to add the 'room' column if it doesn't exist
    try:
        cursor.execute("ALTER TABLE timetable ADD COLUMN room TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()

# Time slots and days
time_slots = [
    "8:30 - 9:20",
    "9:20 - 10:10",
    "10:20 - 11:10",
    "11:10 - 12:00",
    "12:45 - 1:30",
    "1:30 - 2:20",
    "2:20 - 3:00"
]
days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Add entry route for manual timetable entry
@app.route("/add", methods=["GET", "POST"])
def add_entry():
    if request.method == "POST":
        staff_name = request.form["staff_name"]
        department = request.form["department"]
        year = request.form.get("year", "")
        semester = request.form["semester"]
        subject = request.form["subject"]
        room = request.form.get("room", "")
        day = request.form["day"]
        time = request.form["time"]
        db = get_db()
        # Smart Conflict Check:
        # 1. Check if the Staff is already booked at this time
        # 2. Check if the Class (Dept + Year + Sem) is already booked at this time
        # 3. Check if the Room is already booked at this time (if room is specified)
        conflict = db.execute(
            """SELECT * FROM timetable 
               WHERE (staff_name=? AND day=? AND time=?) 
               OR (department=? AND year=? AND semester=? AND day=? AND time=?)
               OR (room != '' AND room=? AND day=? AND time=?)""",
            (staff_name, day, time, department, year, semester, day, time, room, day, time)
        ).fetchone()
        if conflict:
            flash("⚠️ Conflict detected! Staff, Class, or Room is already booked at this time.")
            return redirect(url_for("add_entry"))
        db.execute(
            "INSERT INTO timetable (staff_name, department, year, semester, subject, room, day, time) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (staff_name, department, year, semester, subject, room, day, time)
        )
        db.commit()
        flash("✅ Entry added successfully!")
        return redirect(url_for("dashboard"))
    return render_template("add_timetable.html", days=days, time_slots=time_slots)

# Edit entry route
@app.route("/edit/<int:entry_id>", methods=["GET", "POST"])
def edit_entry(entry_id):
    db = get_db()
    if request.method == "POST":
        staff_name = request.form["staff_name"]
        department = request.form["department"]
        semester = request.form["semester"]
        subject = request.form["subject"]
        day = request.form["day"]
        time = request.form["time"]
        
        # Update the entry
        db.execute(
            "UPDATE timetable SET staff_name=?, department=?, semester=?, subject=?, day=?, time=? WHERE id=?",
            (staff_name, department, semester, subject, day, time, entry_id)
        )
        db.commit()
        flash("✏️ Entry updated successfully!")
        return redirect(url_for("dashboard"))
    
    entry = db.execute("SELECT * FROM timetable WHERE id=?", (entry_id,)).fetchone()
    return render_template("Edit_Timetable.html", entry=entry, days=days, time_slots=time_slots)

# Delete entry route
@app.route("/delete/<int:entry_id>", methods=["POST"])
def delete_entry(entry_id):
    db = get_db()
    db.execute("DELETE FROM timetable WHERE id=?", (entry_id,))
    db.commit()
    flash("🗑️ Entry deleted successfully!")
    return redirect(url_for("dashboard"))

# Auto-assign slot route (from check_slots page)
@app.route("/auto_assign", methods=["POST"])
def auto_assign():
    staff_name = request.form.get("staff_name")
    department = request.form.get("department")
    year = request.form.get("year", "")
    semester = request.form.get("semester", "")
    subject = request.form.get("subject")
    db = get_db()
    used = db.execute(
        "SELECT day, time FROM timetable WHERE staff_name=? AND department=?",
        (staff_name, department)
    ).fetchall()
    used_slots = {(row["day"], row["time"]) for row in used}
    assigned = False
    for t in time_slots:
        for d in days:
            if (d, t) not in used_slots:
                db.execute(
                    "INSERT INTO timetable (staff_name, department, year, semester, subject, day, time) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (staff_name, department, year, semester, subject, d, t)
                )
                db.commit()
                flash(f"Auto-assigned {subject} to {d} {t}.")
                assigned = True
                break
        if assigned:
            break
    if not assigned:
        flash("No free slot available for auto-assign.")
    return redirect(url_for("check_slots"))

# Home route
@app.route("/")
def home():
    return redirect(url_for("dashboard"))

# Dashboard route
@app.route("/dashboard")
def dashboard():
    db = get_db()

    # Filter parameters
    staff_name = request.args.get("staff_name")
    department = request.args.get("department")
    year = request.args.get("year")

    query = "SELECT * FROM timetable WHERE 1=1"
    params = []

    if staff_name:
        query += " AND staff_name = ?"
        params.append(staff_name)
    if department:
        query += " AND department = ?"
        params.append(department)
    if year:
        query += " AND year = ?"
        params.append(year)

    entries = db.execute(query, params).fetchall()

    # Get distinct values for filter buttons
    departments = db.execute("SELECT DISTINCT department FROM timetable WHERE department != '' ORDER BY department").fetchall()
    years = db.execute("SELECT DISTINCT year FROM timetable WHERE year != '' ORDER BY year").fetchall()
    semesters = db.execute("SELECT DISTINCT semester FROM timetable WHERE semester != '' ORDER BY semester").fetchall()
    staff_list = db.execute("SELECT DISTINCT staff_name FROM timetable WHERE staff_name != '' ORDER BY staff_name").fetchall()

    # Analytics Data for Charts
    dept_counts = db.execute("SELECT department, COUNT(*) as count FROM timetable GROUP BY department").fetchall()
    staff_counts = db.execute("SELECT staff_name, COUNT(*) as count FROM timetable GROUP BY staff_name").fetchall()
    
    # Prepare data for frontend (Lists for Chart.js)
    dept_analytics = {"labels": [row["department"] for row in dept_counts], "data": [row["count"] for row in dept_counts]}
    staff_analytics = {"labels": [row["staff_name"] for row in staff_counts], "data": [row["count"] for row in staff_counts]}

    # Logic for Grid View (Department Wise or Staff Wise)
    mode = request.args.get("mode")
    
    # Save current view settings to session for Export functionality
    if mode:
        session['export_mode'] = mode
        session['export_dept'] = request.args.get("department")
        session['export_year'] = request.args.get("year")
        session['export_sem'] = request.args.get("semester")
        session['export_staff'] = request.args.get("staff_name")

    grid_view = False
    schedule = {d: {t: None for t in time_slots} for d in days}
    filter_desc = ""

    if mode == 'dept':
        dept = request.args.get("department")
        yr = request.args.get("year")
        sem = request.args.get("semester")
        if dept and yr and sem:
            grid_view = True
            filter_desc = f"Class Timetable: {dept} - Year {yr} - Semester {sem}"
            rows = db.execute("SELECT * FROM timetable WHERE department=? AND year=? AND semester=?", (dept, yr, sem)).fetchall()
            for row in rows:
                if row["day"] in schedule and row["time"] in schedule[row["day"]]:
                    schedule[row["day"]][row["time"]] = row

    elif mode == 'staff':
        staff = request.args.get("staff_name")
        if staff:
            grid_view = True
            filter_desc = f"Staff Timetable: {staff}"
            rows = db.execute("SELECT * FROM timetable WHERE staff_name=?", (staff,)).fetchall()
            for row in rows:
                if row["day"] in schedule and row["time"] in schedule[row["day"]]:
                    schedule[row["day"]][row["time"]] = row

    return render_template("dashboard.html", entries=entries, departments=departments, years=years, semesters=semesters, staff_list=staff_list, time_slots=time_slots, days=days, grid_view=grid_view, schedule=schedule, filter_desc=filter_desc, username=session.get("username", "Guest"), dept_analytics=dept_analytics, staff_analytics=staff_analytics)

# Export Excel Route
@app.route("/export_excel")
def export_excel():
    db = get_db()
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("Error: 'openpyxl' library is missing. Please install it using 'pip install openpyxl'.")
        return redirect(url_for('dashboard'))

    # Try to get params from URL, otherwise fall back to session
    mode = request.args.get("mode") or session.get('export_mode')
    schedule = {d: {t: "" for t in time_slots} for d in days}
    filename = "timetable.xlsx"

    if mode == 'dept':
        dept = request.args.get("department") or session.get('export_dept')
        yr = request.args.get("year") or session.get('export_year')
        sem = request.args.get("semester") or session.get('export_sem')
        if dept and yr and sem:
            filename = f"{dept}_{yr}_Sem{sem}_timetable.xlsx"
            rows = db.execute("SELECT * FROM timetable WHERE department=? AND year=? AND semester=?", (dept, yr, sem)).fetchall()
            for row in rows:
                if row["day"] in schedule and row["time"] in schedule[row["day"]]:
                    # Format: Subject (Staff Name) [Room]
                    room_str = f" [{row['room']}]" if row['room'] else ""
                    schedule[row["day"]][row["time"]] = f"{row['subject']} ({row['staff_name']}){room_str}"

    elif mode == 'staff':
        staff = request.args.get("staff_name") or session.get('export_staff')
        if staff:
            filename = f"{staff}_timetable.xlsx"
            rows = db.execute("SELECT * FROM timetable WHERE staff_name=?", (staff,)).fetchall()
            for row in rows:
                if row["day"] in schedule and row["time"] in schedule[row["day"]]:
                    # Format: Subject (Dept - Y:Year) [Room]
                    room_str = f" [{row['room']}]" if row['room'] else ""
                    schedule[row["day"]][row["time"]] = f"{row['subject']} ({row['department']} - Y:{row['year']}){room_str}"

    # Generate Excel with openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Write Header Row
    ws.cell(row=1, column=1, value="Day").font = bold_font
    
    col_idx = 2
    col_map = {} # Map slot time to column index
    break_col_idx = -1
    lunch_col_idx = -1

    for slot in time_slots:
        ws.cell(row=1, column=col_idx, value=slot).font = bold_font
        col_map[slot] = col_idx
        col_idx += 1
        
        if slot == "9:20 - 10:10":
            ws.cell(row=1, column=col_idx, value="").font = bold_font
            ws.column_dimensions[get_column_letter(col_idx)].width = 3 # Narrow column
            break_col_idx = col_idx
            col_idx += 1
            
        if slot == "11:10 - 12:00":
            ws.cell(row=1, column=col_idx, value="").font = bold_font
            ws.column_dimensions[get_column_letter(col_idx)].width = 3 # Narrow column
            lunch_col_idx = col_idx
            col_idx += 1

    # Write Data Rows
    break_letters = ["B", "R", "E", "A", "K"]
    lunch_letters = ["L", "U", "N", "C", "H"]

    for row_idx, d in enumerate(days, start=2):
        day_index = row_idx - 2
        ws.cell(row=row_idx, column=1, value=d).alignment = center_align
        for slot in time_slots:
            val = schedule[d][slot] or "-"
            ws.cell(row=row_idx, column=col_map[slot], value=val).alignment = center_align
        
        # Add Break Letter (B, R, E, A, K)
        if break_col_idx != -1 and day_index < len(break_letters):
            cell = ws.cell(row=row_idx, column=break_col_idx, value=break_letters[day_index])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = bold_font

        # Add Lunch Letter (L, U, N, C, H)
        if lunch_col_idx != -1 and day_index < len(lunch_letters):
            cell = ws.cell(row=row_idx, column=lunch_col_idx, value=lunch_letters[day_index])
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = bold_font

    # Apply borders
    end_row = 2 + len(days) - 1
    for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=col_idx-1):
        for cell in row:
            cell.border = border

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    output = make_response(out.getvalue())
    output.headers["Content-Disposition"] = f"attachment; filename={filename}"
    output.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return output

@app.route("/export_pdf")
def export_pdf():
    db = get_db()
    try:
        from xhtml2pdf import pisa
    except ImportError:
        flash("Error: 'xhtml2pdf' library is missing. Please install it using 'pip install xhtml2pdf'.")
        return redirect(url_for('dashboard'))

    mode = request.args.get("mode") or session.get('export_mode')
    schedule = {d: {t: "" for t in time_slots} for d in days}
    filename = "timetable.pdf"
    title = "Timetable"

    if mode == 'dept':
        dept = request.args.get("department") or session.get('export_dept')
        yr = request.args.get("year") or session.get('export_year')
        sem = request.args.get("semester") or session.get('export_sem')
        if dept and yr and sem:
            filename = f"{dept}_{yr}_Sem{sem}_timetable.pdf"
            title = f"Department: {dept} | Year: {yr} | Sem: {sem}"
            rows = db.execute("SELECT * FROM timetable WHERE department=? AND year=? AND semester=?", (dept, yr, sem)).fetchall()
            for row in rows:
                if row["day"] in schedule and row["time"] in schedule[row["day"]]:
                    room_str = f" [{row['room']}]" if row['room'] else ""
                    schedule[row["day"]][row["time"]] = f"{row['subject']} ({row['staff_name']}){room_str}"

    elif mode == 'staff':
        staff = request.args.get("staff_name") or session.get('export_staff')
        if staff:
            filename = f"{staff}_timetable.pdf"
            title = f"Staff: {staff}"
            rows = db.execute("SELECT * FROM timetable WHERE staff_name=?", (staff,)).fetchall()
            for row in rows:
                if row["day"] in schedule and row["time"] in schedule[row["day"]]:
                    room_str = f" [{row['room']}]" if row['room'] else ""
                    schedule[row["day"]][row["time"]] = f"{row['subject']} ({row['department']} - Y:{row['year']}){room_str}"

    # Generate HTML for PDF
    html_content = f"""
    <html>
    <head>
        <style>
            @page {{ size: landscape; margin: 1cm; }}
            body {{ font-family: Helvetica, sans-serif; }}
            h2 {{ text-align: center; color: #333; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ border: 1px solid #444; padding: 6px; text-align: center; font-size: 10px; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            .break-col {{ background-color: #e0e0e0; color: #555; font-weight: bold; vertical-align: middle; width: 15px; }}
            .day {{ background-color: #eee; font-weight: bold; width: 80px; }}
        </style>
    </head>
    <body>
        <h2>{title}</h2>
        <table>
            <thead>
                <tr>
                    <th>Day / Time</th>
                    {''.join([f'<th>{t}</th>' + ('<th class="break-col"></th>' if t == "9:20 - 10:10" else '') + ('<th class="break-col"></th>' if t == "11:10 - 12:00" else '') for t in time_slots])}
                </tr>
            </thead>
            <tbody>
    """
    
    break_letters = ["B", "R", "E", "A", "K"]
    lunch_letters = ["L", "U", "N", "C", "H"]

    for i, d in enumerate(days):
        html_content += f'<tr><td class="day">{d}</td>'
        for t in time_slots:
            html_content += f'<td>{schedule[d][t] or "-"}</td>'
            if t == "9:20 - 10:10":
                html_content += f'<td class="break-col">{break_letters[i] if i < len(break_letters) else ""}</td>'
            if t == "11:10 - 12:00":
                html_content += f'<td class="break-col">{lunch_letters[i] if i < len(lunch_letters) else ""}</td>'
        html_content += '</tr>'

    html_content += """
            </tbody>
        </table>
    </body>
    </html>
    """

    # Create PDF
    pdf = io.BytesIO()
    pisa_status = pisa.CreatePDF(io.StringIO(html_content), dest=pdf)

    if pisa_status.err:
        return "Error generating PDF", 500

    response = make_response(pdf.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

# Check slots route
@app.route("/check_slots", methods=["GET", "POST"])
def check_slots():
    slots = None
    staff_name = None
    department = None
    message = None
    if request.method == "POST":
        staff_name = request.form.get("staff_name")
        department = request.form.get("department")
        year = request.form.get("year", "")
        semester = request.form.get("semester", "")
        auto_assign = request.form.get("auto_assign")
        subject = request.form.get("subject")
        db = get_db()
        used = db.execute(
            "SELECT day, time, staff_name, subject, year, semester, id FROM timetable WHERE staff_name=? AND department=?",
            (staff_name, department)
        ).fetchall()
        used_slots = {(row["day"], row["time"]): row for row in used}
        # Auto-assign logic
        if auto_assign and subject:
            assigned = False
            for t in time_slots:
                for d in days:
                    if (d, t) not in used_slots:
                        db.execute(
                            "INSERT INTO timetable (staff_name, department, year, semester, subject, day, time) VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (staff_name, department, year, semester, subject, d, t)
                        )
                        db.commit()
                        message = f"Auto-assigned {subject} to {d} {t}."
                        assigned = True
                        break
                if assigned:
                    break
            if not assigned:
                message = "No free slot available for auto-assign."
            # Refresh used_slots after auto-assign
            used = db.execute(
                "SELECT day, time, staff_name, subject, year, semester, id FROM timetable WHERE staff_name=? AND department=?",
                (staff_name, department)
            ).fetchall()
            used_slots = {(row["day"], row["time"]): row for row in used}
        # Build a grid: rows=time_slots, columns=days
        timetable_grid = []
        for t in time_slots:
            row = []
            for d in days:
                key = (d, t)
                if key in used_slots:
                    entry = used_slots[key]
                    # Display format: Subject (Y: 1 S: 2)
                    display_text = f"{entry['subject']} (Y:{entry['year']} S:{entry['semester']})"
                    row.append({"subject": display_text, "entry_id": entry['id']})
                else:
                    row.append({"subject": "Free", "entry_id": None})
            timetable_grid.append({"time": t, "slots": row})
        slots = timetable_grid
    return render_template("check_slots.html", slots=slots, staff_name=staff_name, department=department, days=days, message=message)

# Delete slot route (for check_slots page)
@app.route("/delete_slot/<int:entry_id>", methods=["POST"])
def delete_slot(entry_id):
    db = get_db()
    db.execute("DELETE FROM timetable WHERE id=?", (entry_id,))
    db.commit()
    flash("🗑️ Slot deleted successfully!")
    return redirect(url_for("check_slots"))

# Login/logout routes
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        # Simple password check (You can change "admin" to whatever you want)
        if password == "admin":
            session["username"] = username
            flash("Logged in successfully!")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid password!")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.")
    return redirect(url_for("login"))

# Main
if __name__ == "__main__":
    init_db()
    app.run(debug=True, host='0.0.0.0')
